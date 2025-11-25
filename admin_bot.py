"""
ربات مدیریتی برای مدیریت کانال‌ها و مشاهده آمار
"""
import os
import json
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler, JobQueue
from database import Database
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class AdminBot:
    # حالت‌های مکالمه
    WAITING_CHANNEL_USERNAME = 1
    WAITING_CHANNEL_CATEGORY = 2
    WAITING_CHANNEL_TO_REMOVE = 3
    WAITING_CATEGORY_NAME = 4
    
    def __init__(self):
        self.config_file = 'admin_config.json'
        self.db = Database()
        self.config = self.load_config()
        # همگام‌سازی دسته‌بندی‌های موجود از channels به categories
        try:
            synced_count = self.db.sync_categories_from_channels()
            if synced_count > 0:
                print(f"✅ {synced_count} دسته‌بندی از channels به categories همگام‌سازی شد")
        except Exception as e:
            print(f"⚠️ خطا در همگام‌سازی دسته‌بندی‌ها: {e}")
        
    def load_config(self):
        """بارگذاری تنظیمات"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_config(self):
        """ذخیره تنظیمات"""
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
    
    def get_message(self, update: Update):
        """دریافت پیام از update (message یا callback_query.message)"""
        return update.effective_message
    
    def get_main_keyboard(self):
        """کیبورد اصلی"""
        keyboard = [
            [KeyboardButton("📋 لیست کانال‌ها"), KeyboardButton("📊 مشاهده آمار")],
            [KeyboardButton("➕ افزودن کانال"), KeyboardButton("🗑️ حذف کانال")],
            [KeyboardButton("📁 ایجاد دسته‌بندی"), KeyboardButton("🗑️ حذف دسته‌بندی")],
            [KeyboardButton("📋 لیست دسته‌بندی‌ها"), KeyboardButton("📥 خروجی اکسل")],
            [KeyboardButton("⚡ بررسی فوری"), KeyboardButton("🔄 Reset آمار")]
        ]
        return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    def get_inline_keyboard(self):
        """کیبورد اینلاین اصلی"""
        keyboard = [
            [
                InlineKeyboardButton("📋 لیست کانال‌ها", callback_data="list_channels"),
                InlineKeyboardButton("📊 آمار", callback_data="show_stats")
            ],
            [
                InlineKeyboardButton("➕ افزودن کانال", callback_data="add_channel"),
                InlineKeyboardButton("🗑️ حذف کانال", callback_data="remove_channel")
            ],
            [
                InlineKeyboardButton("📥 خروجی اکسل", callback_data="export_excel"),
                InlineKeyboardButton("⚡ بررسی فوری", callback_data="trigger_check")
            ]
        ]
        return InlineKeyboardMarkup(keyboard)
    
    def trigger_immediate_check(self, user_id: int = None) -> bool:
        """ایجاد فایل flag برای بررسی فوری کانال‌ها"""
        try:
            # استفاده از دایرکتوری data برای فایل‌های flag (مشترک بین containers)
            data_dir = os.path.join(os.getcwd(), 'data')
            os.makedirs(data_dir, exist_ok=True)
            
            flag_file = 'trigger_check.flag'
            flag_path = os.path.join(data_dir, flag_file)
            with open(flag_path, 'w') as f:
                if user_id:
                    f.write(str(user_id))  # ذخیره user_id در فایل
                else:
                    f.write('')  # ایجاد فایل خالی
            return True
        except Exception as e:
            print(f"خطا در ایجاد فایل flag: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def trigger_join_channel(self, channel_id: int, channel_identifier: str):
        """ایجاد فایل flag برای join کردن کانال"""
        try:
            # استفاده از دایرکتوری data برای فایل‌های flag (مشترک بین containers)
            data_dir = os.path.join(os.getcwd(), 'data')
            os.makedirs(data_dir, exist_ok=True)
            
            join_file = 'join_channel.flag'
            join_path = os.path.join(data_dir, join_file)
            join_data = {
                'channel_id': channel_id,
                'channel_identifier': channel_identifier
            }
            with open(join_path, 'w', encoding='utf-8') as f:
                json.dump(join_data, f, ensure_ascii=False, indent=2)
            print(f"🚩 فایل join_channel.flag ایجاد شد در {join_path} برای channel_id: {channel_id}")
            return True
        except Exception as e:
            print(f"خطا در ایجاد فایل join flag: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def trigger_leave_channel(self, channel_id: int, username: str):
        """ایجاد فایل flag برای leave کردن کانال فوری"""
        try:
            # استفاده از دایرکتوری data برای فایل‌های flag (مشترک بین containers)
            data_dir = os.path.join(os.getcwd(), 'data')
            os.makedirs(data_dir, exist_ok=True)
            
            leave_file = 'leave_channel.flag'
            leave_path = os.path.join(data_dir, leave_file)
            leave_data = {
                'channel_id': channel_id,
                'username': username
            }
            with open(leave_path, 'w', encoding='utf-8') as f:
                json.dump(leave_data, f, ensure_ascii=False, indent=2)
            print(f"🚩 فایل leave_channel.flag ایجاد شد در {leave_path} برای channel_id: {channel_id}")
            return True
        except Exception as e:
            print(f"خطا در ایجاد فایل leave flag: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def read_notification_file(self) -> dict:
        """خواندن فایل notification بدون حذف"""
        notification_file = 'check_notification.json'
        notification_path = os.path.join(os.getcwd(), notification_file)
        
        if not os.path.exists(notification_path):
            return None
        
        try:
            with open(notification_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        except Exception as e:
            print(f"❌ خطا در خواندن فایل notification: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def delete_notification_file(self):
        """حذف فایل notification"""
        notification_file = 'check_notification.json'
        notification_path = os.path.join(os.getcwd(), notification_file)
        if os.path.exists(notification_path):
            try:
                os.remove(notification_path)
                print(f"🗑️ فایل notification حذف شد")
            except Exception as e:
                print(f"⚠️ خطا در حذف فایل notification: {e}")
    
    async def check_and_notify(self, context: ContextTypes.DEFAULT_TYPE):
        """چک کردن فایل notification و اطلاع‌رسانی به کاربر"""
        try:
            # خواندن فایل notification (بدون حذف)
            notification = self.read_notification_file()
            
            if notification:
                user_id = notification.get('user_id')
                channels_count = notification.get('channels_count', 0)
                success = notification.get('success', False)
                
                print(f"📩 فایل notification پیدا شد!")
                print(f"   user_id={user_id}, channels={channels_count}, success={success}")
                
                if user_id and success:
                    try:
                        # دریافت آخرین زمان آپدیت
                        latest_stats = self.db.get_all_stats()
                        last_update_text = ""
                        if latest_stats:
                            last_stat = latest_stats[0]
                            if last_stat.get('recorded_at'):
                                try:
                                    date_obj = datetime.fromisoformat(last_stat['recorded_at'].replace('Z', '+00:00'))
                                    last_update_text = f"\n🕐 آخرین آپدیت: {date_obj.strftime('%Y-%m-%d %H:%M')}"
                                except:
                                    pass
                        
                        message_text = (
                            "✅ بررسی کانال‌ها تکمیل شد!\n\n"
                            f"📊 تعداد کانال‌های بررسی شده: {channels_count}{last_update_text}\n\n"
                            "حالا می‌توانید:\n"
                            "• آمار را مشاهده کنید (📊 مشاهده آمار)\n"
                            "• خروجی اکسل بگیرید (📥 خروجی اکسل)"
                        )
                        
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=message_text,
                            reply_markup=self.get_main_keyboard()
                        )
                        print(f"✅ پیام اطلاع‌رسانی به کاربر {user_id} با موفقیت ارسال شد!")
                        # فقط بعد از ارسال موفق پیام، فایل را حذف می‌کنیم
                        self.delete_notification_file()
                    except Exception as e:
                        print(f"❌ خطا در ارسال پیام به کاربر {user_id}: {e}")
                        import traceback
                        traceback.print_exc()
                elif not user_id:
                    print(f"⚠️ user_id در notification موجود نیست. notification: {notification}")
                    self.delete_notification_file()  # حذف فایل نامعتبر
                elif not success:
                    print(f"⚠️ بررسی موفق نبوده است. notification: {notification}")
                    self.delete_notification_file()  # حذف فایل نامعتبر
        except Exception as e:
            print(f"❌ خطا در check_and_notify: {e}")
            import traceback
            traceback.print_exc()
    
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """دستور /start"""
        user_id = update.effective_user.id
        username = update.effective_user.username or ""
        
        message = self.get_message(update)
        if not message:
            return
        
        # افزودن کاربر اول به عنوان ادمین
        if not self.db.is_admin(user_id):
            self.db.add_admin(user_id, username)
            message_text = (
                "✅ به عنوان ادمین اضافه شدید!\n\n"
                "🤖 ربات مدیریت کانال‌های تلگرام\n\n"
                "از دکمه‌های زیر برای مدیریت استفاده کنید:"
            )
        else:
            message_text = (
                "🤖 ربات مدیریت کانال‌های تلگرام\n\n"
                "از دکمه‌های زیر برای مدیریت استفاده کنید:"
            )
        
        await message.reply_text(
            message_text,
            reply_markup=self.get_main_keyboard()
        )
    
    async def check_admin(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
        """بررسی ادمین بودن کاربر"""
        user_id = update.effective_user.id
        if not self.db.is_admin(user_id):
            message = self.get_message(update)
            if message:
                await message.reply_text("❌ شما دسترسی ادمین ندارید!")
            return False
        return True
    
    async def list_channels(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """لیست کانال‌ها"""
        if not await self.check_admin(update, context):
            return
        
        channels = self.db.get_all_active_channels()  # همه کانال‌های فعال را بگیر
        
        message = self.get_message(update)
        if not message:
            return
            
        if not channels:
            await message.reply_text(
                "📭 هیچ کانال فعالی وجود ندارد.\n\n"
                "از دکمه ➕ افزودن کانال برای افزودن کانال جدید استفاده کنید.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        text = "📋 لیست کانال‌ها:\n\n"
        for i, channel in enumerate(channels, 1):
            title = channel.get('title', 'بدون عنوان')
            username = channel['username']
            category = channel.get('category', '')
            is_member = channel.get('is_member', 0)
            member_status = "✅ عضو" if is_member else "⏳ در انتظار عضویت"
            
            text += f"{i}. {title}\n"
            if username.startswith('http') or username.startswith('+'):
                text += f"   🔗 {username[:40]}...\n"
            else:
                text += f"   @{username}\n"
            if category:
                text += f"   📁 {category}\n"
            text += f"   {member_status}\n\n"
        
        text += f"\n📊 تعداد کل: {len(channels)} کانال"
        
        await message.reply_text(text, reply_markup=self.get_main_keyboard())
    
    async def add_channel_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """شروع فرآیند افزودن کانال"""
        if not await self.check_admin(update, context):
            return ConversationHandler.END
        
        message = self.get_message(update)
        if not message:
            return ConversationHandler.END
        
        await message.reply_text(
            "➕ افزودن کانال جدید\n\n"
            "لطفاً یوزرنیم یا لینک کانال را ارسال کنید:\n"
            "• یوزرنیم: channel_name یا @channel_name\n"
            "• لینک پرایوت: https://t.me/+ABC123...\n\n"
            "❌ برای لغو /cancel را بفرستید",
            reply_markup=ReplyKeyboardRemove()
        )
        
        return self.WAITING_CHANNEL_USERNAME
    
    async def add_channel_process(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """پردازش افزودن کانال - مرحله اول: دریافت username"""
        message = self.get_message(update)
        if not message or not message.text:
            return self.WAITING_CHANNEL_USERNAME
        
        input_text = message.text.strip()
        
        # اعتبارسنجی: بررسی اینکه آیا این یک دکمه کیبورد است یا خیر
        # دکمه‌های کیبورد اصلی را نادیده بگیر
        main_keyboard_texts = [
            "📋 لیست کانال‌ها", "📊 مشاهده آمار", "➕ افزودن کانال", 
            "🗑️ حذف کانال", "📁 ایجاد دسته‌بندی", "🗑️ حذف دسته‌بندی",
            "📋 لیست دسته‌بندی‌ها", "📥 خروجی اکسل", "⚡ بررسی فوری", "🔄 Reset آمار"
        ]
        
        if input_text in main_keyboard_texts:
            await message.reply_text(
                "⚠️ لطفاً یوزرنیم یا لینک کانال را وارد کنید، نه دکمه کیبورد!\n\n"
                "مثال:\n"
                "• channel_name\n"
                "• @channel_name\n"
                "• https://t.me/+ABC123...",
                reply_markup=self.get_main_keyboard()
            )
            return ConversationHandler.END
        
        # اعتبارسنجی: username باید معتبر باشد
        is_invite_link = input_text.startswith('http') or input_text.startswith('t.me/+') or input_text.startswith('+')
        if not is_invite_link:
            # بررسی اینکه username معتبر است
            username_clean = input_text.lstrip('@').strip()
            # username باید حداقل 5 کاراکتر باشد و فقط شامل حروف، اعداد و _ باشد
            import re
            if not re.match(r'^[a-zA-Z0-9_]{5,}$', username_clean):
                await message.reply_text(
                    "⚠️ یوزرنیم معتبر نیست!\n\n"
                    "یوزرنیم باید:\n"
                    "• حداقل 5 کاراکتر باشد\n"
                    "• فقط شامل حروف انگلیسی، اعداد و _ باشد\n\n"
                    "مثال: channel_name یا @channel_name",
                    reply_markup=self.get_main_keyboard()
                )
                return ConversationHandler.END
        
        # ذخیره username در context
        context.user_data['channel_input'] = input_text
        
        # دریافت لیست دسته‌بندی‌های موجود از دیتابیس
        categories = self.db.get_all_categories()
        
        print(f"DEBUG: دسته‌بندی‌های دریافت شده: {categories}")  # لاگ برای دیباگ
        
        # ساخت کیبورد برای انتخاب دسته‌بندی
        if categories:
            # اگر دسته‌بندی وجود دارد، آن‌ها را نمایش می‌دهیم
            keyboard = []
            for i in range(0, len(categories), 2):
                row = []
                row.append(KeyboardButton(categories[i]))
                if i + 1 < len(categories):
                    row.append(KeyboardButton(categories[i + 1]))
                keyboard.append(row)
            
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)
            category_text = "\n\n📋 دسته‌بندی‌های موجود:\n" + "\n".join([f"• {cat}" for cat in categories])
            
            await message.reply_text(
                "✅ کانال دریافت شد!\n\n"
                "📁 حالا لطفاً دسته‌بندی کانال را انتخاب کنید:" + category_text,
                reply_markup=reply_markup
            )
        else:
            # اگر دسته‌بندی وجود ندارد، پیام راهنما بدون کیبورد
            await message.reply_text(
                "✅ کانال دریافت شد!\n\n"
                "📁 حالا لطفاً دسته‌بندی کانال را وارد کنید:\n\n"
                "💡 هیچ دسته‌بندی موجودی نیست.\n"
                "می‌توانید یک دسته‌بندی جدید وارد کنید (مثلاً: انیمه، خبری، ورزشی و ...)",
                reply_markup=ReplyKeyboardRemove()
            )
        
        return self.WAITING_CHANNEL_CATEGORY
    
    async def add_channel_category(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """پردازش افزودن کانال - مرحله دوم: دریافت دسته‌بندی"""
        message = self.get_message(update)
        if not message or not message.text:
            return self.WAITING_CHANNEL_CATEGORY
        
        category = message.text.strip()
        input_text = context.user_data.get('channel_input')
        user_id = update.effective_user.id
        
        if not input_text:
            await message.reply_text(
                "❌ خطا! لطفاً دوباره از /add شروع کنید.",
                reply_markup=self.get_main_keyboard()
            )
            return ConversationHandler.END
        
        # تشخیص اینکه آیا invite link است یا username
        is_invite_link = input_text.startswith('http') or input_text.startswith('t.me/+') or input_text.startswith('+')
        
        # افزودن کانال با دسته‌بندی (اول بدون is_member)
        if is_invite_link:
            success = self.db.add_channel(input_text, "", user_id, invite_link=input_text, category=category)
            channel_display = f"لینک کانال"
            channel_identifier = input_text
        else:
            username = input_text.lstrip('@')
            success = self.db.add_channel(username, "", user_id, category=category)
            channel_display = f"@{username}"
            channel_identifier = username
        
        # پاک کردن داده‌های موقت
        context.user_data.pop('channel_input', None)
        
        if success:
            # دریافت channel_id اضافه شده
            # برای invite link، از input_text استفاده می‌کنیم (همان چیزی که در دیتابیس ذخیره شده)
            # برای username، از channel_identifier استفاده می‌کنیم
            search_username = input_text if is_invite_link else channel_identifier
            channel_info = self.db.get_channel_by_username(search_username)
            
            if channel_info:
                channel_id = channel_info['id']
                # ایجاد فایل flag برای join کردن کانال
                self.trigger_join_channel(channel_id, channel_identifier)
            else:
                # اگر نتوانستیم channel_info را پیدا کنیم، سعی می‌کنیم از آخرین کانال اضافه شده استفاده کنیم
                print(f"⚠️ نتوانستیم channel_info را برای {search_username} پیدا کنیم")
            
            await message.reply_text(
                f"✅ کانال {channel_display} با دسته‌بندی '{category}' با موفقیت اضافه شد!\n\n"
                f"🔄 در حال پیوستن به کانال...\n"
                f"ربات رصد به زودی به کانال خواهد پیوست.",
                reply_markup=self.get_main_keyboard()
            )
        else:
            await message.reply_text(
                f"❌ خطا! کانال {channel_display} قبلاً اضافه شده یا خطایی رخ داد.\n"
                "لطفاً دوباره امتحان کنید.",
                reply_markup=self.get_main_keyboard()
            )
        
        return ConversationHandler.END
    
    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """لغو عملیات"""
        message = self.get_message(update)
        if message:
            await message.reply_text(
                "❌ عملیات لغو شد.",
                reply_markup=self.get_main_keyboard()
            )
        return ConversationHandler.END
    
    async def create_category_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """شروع فرآیند ایجاد دسته‌بندی جدید"""
        if not await self.check_admin(update, context):
            return ConversationHandler.END
        
        message = self.get_message(update)
        if not message:
            return ConversationHandler.END
        
        await message.reply_text(
            "📁 ایجاد دسته‌بندی جدید\n\n"
            "لطفاً نام دسته‌بندی جدید را ارسال کنید:\n"
            "مثال: تکنولوژی\n\n"
            "❌ برای لغو /cancel را بفرستید",
            reply_markup=ReplyKeyboardRemove()
        )
        
        return self.WAITING_CATEGORY_NAME
    
    async def create_category_process(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """پردازش ایجاد دسته‌بندی"""
        message = self.get_message(update)
        if not message or not message.text:
            return self.WAITING_CATEGORY_NAME
        
        category_name = message.text.strip()
        
        # بررسی اینکه دسته‌بندی قبلاً وجود داشته
        existing_categories = self.db.get_all_categories()
        if category_name in existing_categories:
            await message.reply_text(
                f"⚠️ دسته‌بندی '{category_name}' قبلاً وجود دارد!",
                reply_markup=self.get_main_keyboard()
            )
            return ConversationHandler.END
        
        # ذخیره دسته‌بندی در جدول categories
        success = self.db.add_category(category_name)
        
        if success:
            await message.reply_text(
                f"✅ دسته‌بندی '{category_name}' با موفقیت ایجاد شد!\n\n"
                f"از این به بعد می‌توانید هنگام افزودن کانال، این دسته‌بندی را انتخاب کنید.",
                reply_markup=self.get_main_keyboard()
            )
        else:
            await message.reply_text(
                f"❌ خطا در ایجاد دسته‌بندی '{category_name}'!",
                reply_markup=self.get_main_keyboard()
            )
        
        return ConversationHandler.END
    
    async def list_categories(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """لیست دسته‌بندی‌ها"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message:
            return
        
        # دریافت لیست دسته‌بندی‌های موجود
        categories = self.db.get_all_categories()
        
        if not categories:
            await message.reply_text(
                "📭 هیچ دسته‌بندی موجودی وجود ندارد.\n\n"
                "می‌توانید با دکمه 📁 ایجاد دسته‌بندی، دسته‌بندی جدید ایجاد کنید.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        text = "📁 لیست دسته‌بندی‌ها:\n\n"
        
        for i, category in enumerate(sorted(categories), 1):
            # شمارش تعداد کانال‌های این دسته
            channel_count = self.db.get_channels_count_by_category(category)
            text += f"{i}. {category}\n   📊 تعداد کانال‌ها: {channel_count}\n\n"
        
        text += f"\n📊 تعداد کل: {len(categories)} دسته‌بندی"
        
        await message.reply_text(text, reply_markup=self.get_main_keyboard())
    
    async def delete_category_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """شروع فرآیند حذف دسته‌بندی"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message:
            return
        
        # دریافت لیست دسته‌بندی‌های موجود
        categories = self.db.get_all_categories()
        
        if not categories:
            await message.reply_text(
                "📭 هیچ دسته‌بندی موجودی برای حذف وجود ندارد.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        # ساخت کیبورد اینلاین برای انتخاب دسته‌بندی
        keyboard = []
        for category in sorted(categories):
            # شمارش تعداد کانال‌های این دسته
            channel_count = self.db.get_channels_count_by_category(category)
            button_text = f"📁 {category} ({channel_count} کانال)"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=f"delete_category_{category}")])
        
        keyboard.append([InlineKeyboardButton("❌ لغو", callback_data="cancel_delete_category")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await message.reply_text(
            "🗑️ حذف دسته‌بندی\n\n"
            "لطفاً دسته‌بندی مورد نظر برای حذف را انتخاب کنید:\n\n"
            "⚠️ توجه: با حذف دسته‌بندی، دسته‌بندی همه کانال‌های آن به 'بدون دسته‌بندی' تغییر می‌کند.",
            reply_markup=reply_markup
        )
    
    async def remove_channel_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """شروع فرآیند حذف کانال"""
        if not await self.check_admin(update, context):
            return
        
        channels = self.db.get_all_active_channels()  # همه کانال‌های فعال را نمایش می‌دهیم
        
        message = self.get_message(update)
        if not message:
            return
        
        if not channels:
            await message.reply_text(
                "📭 هیچ کانال فعالی برای حذف وجود ندارد.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        # ساخت کیبورد اینلاین برای انتخاب کانال
        keyboard = []
        for channel in channels[:20]:  # حداکثر 20 کانال
            title = channel.get('title', 'بدون عنوان')[:30]  # محدود کردن طول
            username = channel['username']
            button_text = f"{title} (@{username})"
            keyboard.append([InlineKeyboardButton(button_text, callback_data=f"remove_{channel['id']}")])
        
        keyboard.append([InlineKeyboardButton("❌ لغو", callback_data="cancel_remove")])
        
        await message.reply_text(
            "🗑️ حذف کانال\n\n"
            "لطفاً کانالی که می‌خواهید حذف کنید را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    async def remove_channel_confirm(self, update: Update, context: ContextTypes.DEFAULT_TYPE, channel_id: int):
        """تأیید حذف کانال"""
        query = update.callback_query
        await query.answer()
        
        channel = self.db.get_channel_by_id(channel_id)
        
        if channel:
            username = channel['username']
            channel_id = channel['id']
            success = self.db.remove_channel(username)
            
            if success:
                # ایجاد فایل flag برای خروج فوری از کانال
                self.trigger_leave_channel(channel_id, username)
                
                await query.edit_message_text(
                    f"✅ کانال @{username} با موفقیت حذف شد!\n\n"
                    f"🔄 ربات در حال خروج از کانال است...",
                    reply_markup=self.get_inline_keyboard()
                )
            else:
                await query.edit_message_text(
                    f"❌ خطا در حذف کانال!",
                    reply_markup=self.get_inline_keyboard()
                )
        else:
            await query.edit_message_text(
                "❌ کانال یافت نشد!",
                reply_markup=self.get_inline_keyboard()
            )
    
    async def show_stats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """نمایش آمار"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message:
            return
        
        stats = self.db.get_all_stats()
        
        if not stats:
            await message.reply_text(
                "📊 هیچ آماری ثبت نشده است.\n\n"
                "ربات رصد باید در حال اجرا باشد تا آمار ثبت شود.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        # گروه‌بندی آمار بر اساس دسته‌بندی
        stats_by_category = {}
        for stat in stats:
            category = stat.get('category') or 'بدون دسته‌بندی'
            if category not in stats_by_category:
                stats_by_category[category] = []
            stats_by_category[category].append(stat)
        
        text = "📊 آمار کانال‌ها\n"
        text += "─" * 30 + "\n\n"
        
        total_channels = sum(len(stats) for stats in stats_by_category.values())
        text += f"📈 تعداد کل کانال‌ها: {total_channels}\n"
        text += f"📁 تعداد دسته‌بندی‌ها: {len(stats_by_category)}\n\n"
        
        for category, category_stats in sorted(stats_by_category.items()):
            text += f"┌─ 📁 {category}\n"
            text += f"│  تعداد: {len(category_stats)} کانال\n"
            text += f"├" + "─" * 28 + "\n"
            
            for i, stat in enumerate(category_stats, 1):
                channel_id = stat.get('id')
                username = stat.get('username', 'نامشخص')
                title = stat.get('title', 'بدون عنوان')
                member_count = stat.get('member_count', 0) or 0
                member_change = stat.get('member_change', 0) or 0
                recorded_at = stat.get('recorded_at', '')
                views_count = stat.get('views_count', 0) or 0
                views_change = stat.get('views_change', 0) or 0
                
                # محاسبه تغییر نسبت به دیروز
                change_from_yesterday = member_change
                change_percent = 0
                if member_count > 0 and member_change != 0:
                    change_percent = (member_change / (member_count - member_change)) * 100 if (member_count - member_change) > 0 else 0
                
                # محاسبه تغییر نسبت به اولین روز
                change_from_first = 0
                if channel_id:
                    first_stats = self.db.get_first_stats(channel_id)
                    if first_stats and first_stats.get('member_count'):
                        change_from_first = member_count - (first_stats.get('member_count', 0) or 0)
                
                # نمایش username یا invite link
                if username.startswith('http') or username.startswith('+') or username.startswith('t.me'):
                    username_display = "🔗 لینک پرایوت"
                elif not username or username == 'نامشخص':
                    username_display = "❓ نامشخص"
                else:
                    username_display = f"@{username}"
                
                # فرمت تاریخ آخرین آپدیت
                last_update_text = ""
                if recorded_at:
                    try:
                        date_obj = datetime.fromisoformat(recorded_at.replace('Z', '+00:00'))
                        last_update_text = date_obj.strftime('%Y/%m/%d %H:%M')
                    except:
                        last_update_text = recorded_at[:16] if len(recorded_at) >= 16 else recorded_at
                
                # تعیین وضعیت تغییر
                status_icon = "📈" if member_change > 0 else "📉" if member_change < 0 else "➡️"
                
                # نمایش اطلاعات کانال
                text += f"│\n"
                text += f"│ {i}. {title}\n"
                text += f"│    {username_display}\n"
                text += f"│    👥 اعضا: {member_count:,}\n"
                
                # تغییرات
                if change_from_yesterday != 0:
                    change_sign = "+" if change_from_yesterday > 0 else ""
                    text += f"│    {status_icon} تغییر امروز: {change_sign}{change_from_yesterday:,}"
                    if abs(change_percent) > 0.01:
                        text += f" ({change_sign}{abs(change_percent):.1f}%)"
                    text += "\n"
                
                if change_from_first != 0:
                    first_sign = "+" if change_from_first > 0 else ""
                    text += f"│    📊 تغییر از ابتدا: {first_sign}{change_from_first:,}\n"
                
                # نمایش تاریخ آخرین آپدیت
                if last_update_text:
                    text += f"│    🕐 آخرین آپدیت: {last_update_text}\n"
                
                # نمایش views اگر موجود باشد
                if views_count > 0:
                    text += f"│    👁️ بازدید: {views_count:,}\n"
                
                # خط جداکننده بین کانال‌ها (جز آخرین)
                if i < len(category_stats):
                    text += f"│\n"
            
            text += f"└" + "─" * 28 + "\n\n"
        
        # اگر متن خیلی طولانی شد (بیش از 4000 کاراکتر)، فقط خلاصه نشان می‌دهیم
        if len(text) > 4000:
            # ساخت یک خلاصه کوتاه‌تر
            summary_text = "📊 آمار کانال‌ها\n"
            summary_text += "─" * 30 + "\n\n"
            summary_text += f"📈 تعداد کل کانال‌ها: {total_channels}\n"
            summary_text += f"📁 تعداد دسته‌بندی‌ها: {len(stats_by_category)}\n\n"
            summary_text += "⚠️ متن کامل خیلی طولانی است.\n"
            summary_text += "لطفاً از دکمه 📥 خروجی اکسل برای دریافت اطلاعات کامل استفاده کنید.\n\n"
            summary_text += "📊 خلاصه:\n"
            
            for category, category_stats in sorted(stats_by_category.items()):
                summary_text += f"📁 {category}: {len(category_stats)} کانال\n"
            
            await message.reply_text(summary_text, reply_markup=self.get_main_keyboard())
        else:
            await message.reply_text(text, reply_markup=self.get_main_keyboard())
    
    async def reset_stats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Reset آمار کانال‌ها"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message:
            return
        
        # ساخت کیبورد اینلاین برای تایید
        keyboard = [
            [InlineKeyboardButton("✅ بله، Reset کن", callback_data="confirm_reset_all")],
            [InlineKeyboardButton("❌ لغو", callback_data="cancel_reset")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await message.reply_text(
            "🔄 Reset آمار\n\n"
            "⚠️ آیا مطمئن هستید که می‌خواهید همه آمار شمارش (تغییرات) را صفر کنید؟\n\n"
            "این کار:\n"
            "✅ تغییرات (member_change, views_change, ...) را صفر می‌کند\n"
            "✅ تعداد فعلی اعضا (member_count) را حفظ می‌کند\n"
            "✅ عنوان و اطلاعات کانال را حفظ می‌کند\n\n"
            "این عملیات قابل بازگشت نیست!",
            reply_markup=reply_markup
        )
    
    def create_excel(self) -> str:
        """ایجاد فایل اکسل - گروه‌بندی بر اساس دسته‌بندی"""
        stats = self.db.get_all_stats()
        
        # ایجاد workbook
        wb = openpyxl.Workbook()
        
        # حذف sheet پیش‌فرض
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # هدرها
        headers = ["Title", "Username Date", "Member Count", "Change from Yesterday", "Change from First Day", "Positive Change"]
        
        # استایل هدر
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # گروه‌بندی آمار بر اساس دسته‌بندی
        stats_by_category = {}
        for stat in stats:
            category = stat.get('category') or 'بدون دسته‌بندی'
            if category not in stats_by_category:
                stats_by_category[category] = []
            stats_by_category[category].append(stat)
        
        # ایجاد sheet برای هر دسته‌بندی
        for category, category_stats in stats_by_category.items():
            ws = wb.create_sheet(title=category[:31])  # محدودیت 31 کاراکتر برای نام sheet
            
            # اضافه کردن هدر
            ws.append(headers)
            
            # استایل هدر
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # داده‌ها برای این دسته‌بندی
            for stat in category_stats:
                channel_id = stat.get('id')
                username = stat.get('username', '')
                title = stat.get('title', 'بدون عنوان')
                recorded_at = stat.get('recorded_at', '')
                
                # فرمت تاریخ
                if recorded_at:
                    try:
                        date_obj = datetime.fromisoformat(recorded_at.replace('Z', '+00:00'))
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except:
                        date_str = recorded_at[:10] if len(recorded_at) >= 10 else datetime.now().strftime('%Y-%m-%d')
                else:
                    date_str = datetime.now().strftime('%Y-%m-%d')
                
                member_count = stat.get('member_count', 0) or 0
                
                # محاسبه تغییر نسبت به دیروز
                change_from_yesterday = 0
                if channel_id:
                    yesterday_stats = self.db.get_yesterday_stats(channel_id)
                    if yesterday_stats:
                        change_from_yesterday = member_count - (yesterday_stats.get('member_count', 0) or 0)
                
                # محاسبه تغییر نسبت به اولین روز
                change_from_first_day = 0
                if channel_id:
                    first_stats = self.db.get_first_stats(channel_id)
                    if first_stats:
                        change_from_first_day = member_count - (first_stats.get('member_count', 0) or 0)
                
                positive_change = stat.get('positive_change', 0) or 0
                
                # فرمت username_date مطابق با تصویر: "username YYYY-MM-D"
                # حذف صفر از ابتدای روز اگر یک رقمی باشد (مثلاً 2025-11-1 به جای 2025-11-01)
                date_parts = date_str.split('-')
                if len(date_parts) == 3:
                    day = str(int(date_parts[2]))  # حذف صفر از ابتدای روز
                    date_formatted = f"{date_parts[0]}-{date_parts[1]}-{day}"
                else:
                    date_formatted = date_str
                
                # نمایش previous_telegram_id اگر موجود باشد
                previous_id = stat.get('previous_telegram_id')
                if previous_id:
                    username_date = f"{username} {date_formatted} (قبلی: {previous_id})"
                else:
                    username_date = f"{username} {date_formatted}"
                
                row = [
                    title,
                    username_date,
                    member_count,
                    change_from_yesterday,
                    change_from_first_day,
                    positive_change
                ]
                ws.append(row)
            
            # تنظیم عرض ستون‌ها
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
        
        # ذخیره فایل
        filename = f"channel_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    async def export_excel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """خروجی اکسل - منوی انتخاب نوع خروجی"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message:
            return
        
        # دریافت لیست همه دسته‌بندی‌های موجود
        all_categories = self.db.get_all_categories()
        # دریافت دسته‌بندی‌هایی که کانال فعال دارند
        categories_with_channels = set(self.db.get_categories_with_active_channels())
        
        if not all_categories:
            await message.reply_text(
                "❌ هیچ دسته‌بندی موجودی برای خروجی وجود ندارد!\n\n"
                "لطفاً ابتدا دسته‌بندی ایجاد کنید.",
                reply_markup=self.get_main_keyboard()
            )
            return
        
        categories = all_categories
        
        # ساخت کیبورد اینلاین
        keyboard = [
            [InlineKeyboardButton("📊 خروجی همه دسته‌ها", callback_data="export_all_categories")]
        ]
        
        # اضافه کردن دکمه‌های دسته‌بندی‌ها
        for category in sorted(categories):
            keyboard.append([InlineKeyboardButton(f"📁 {category}", callback_data=f"export_category_{category}")])
        
        keyboard.append([InlineKeyboardButton("❌ لغو", callback_data="cancel_export")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await message.reply_text(
            "📥 انتخاب نوع خروجی:\n\n"
            "لطفاً نوع خروجی مورد نظر را انتخاب کنید:",
            reply_markup=reply_markup
        )
    
    async def export_excel_all(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """خروجی اکسل - همه دسته‌بندی‌ها (فایل‌های جداگانه)"""
        query = update.callback_query
        if query:
            await query.answer("در حال ساخت فایل‌های اکسل...")
            message = query.message
        else:
            message = self.get_message(update)
        
        if not message:
            return
        
        try:
            await message.edit_text("⏳ در حال ساخت فایل‌های اکسل (همه دسته‌بندی‌ها)...\nلطفاً صبر کنید...")
            
            # دریافت لیست همه دسته‌بندی‌ها
            categories = self.db.get_all_categories()
            
            if not categories:
                await message.edit_text(
                    "❌ هیچ دسته‌بندی موجودی برای خروجی وجود ندارد!",
                    reply_markup=self.get_inline_keyboard()
                )
                return
            
            # ساخت فایل برای هر دسته‌بندی
            filenames = []
            for category in sorted(categories):
                filename = self.create_excel_by_category(category)
                filenames.append((filename, category))
            
            # ارسال فایل‌ها
            for filename, category in filenames:
                with open(filename, 'rb') as f:
                    await message.reply_document(
                        document=f,
                        filename=filename,
                        caption=f"📊 خروجی اکسل - دسته‌بندی: {category}"
                    )
                # حذف فایل موقت
                os.remove(filename)
            
            await message.reply_text(
                f"✅ {len(filenames)} فایل اکسل با موفقیت ارسال شد!",
                reply_markup=self.get_main_keyboard()
            )
            
        except Exception as e:
            await message.reply_text(
                f"❌ خطا در ساخت فایل اکسل: {e}",
                reply_markup=self.get_main_keyboard()
            )
            import traceback
            traceback.print_exc()
    
    async def export_excel_category(self, update: Update, context: ContextTypes.DEFAULT_TYPE, category: str):
        """خروجی اکسل - یک دسته‌بندی خاص"""
        query = update.callback_query
        if query:
            await query.answer("در حال ساخت فایل اکسل...")
            message = query.message
        else:
            message = self.get_message(update)
        
        if not message:
            return
        
        try:
            await message.edit_text(f"⏳ در حال ساخت فایل اکسل (دسته‌بندی: {category})...\nلطفاً صبر کنید...")
            filename = self.create_excel_by_category(category)
            
            with open(filename, 'rb') as f:
                await message.reply_document(
                    document=f,
                    filename=filename,
                    caption=f"📊 خروجی اکسل آمار کانال‌ها (دسته‌بندی: {category})"
                )
            
            # حذف فایل موقت
            os.remove(filename)
            
        except Exception as e:
            await message.reply_text(
                f"❌ خطا در ساخت فایل اکسل: {e}",
                reply_markup=self.get_main_keyboard()
            )
    
    def create_excel_by_category(self, category: str) -> str:
        """ایجاد فایل اکسل برای یک دسته‌بندی خاص"""
        stats = self.db.get_all_stats()
        
        # فیلتر کردن آمار بر اساس دسته‌بندی
        category_stats = [stat for stat in stats if stat.get('category') == category]
        
        # ایجاد workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = category[:31]  # محدودیت 31 کاراکتر
        
        # هدرها
        headers = ["Title", "Username Date", "Member Count", "Change from Yesterday", "Change from First Day", "Positive Change"]
        ws.append(headers)
        
        # استایل هدر
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # داده‌ها
        for stat in category_stats:
            channel_id = stat.get('id')
            username = stat.get('username', '')
            title = stat.get('title', 'بدون عنوان')
            recorded_at = stat.get('recorded_at', '')
            
            # فرمت تاریخ
            if recorded_at:
                try:
                    date_obj = datetime.fromisoformat(recorded_at.replace('Z', '+00:00'))
                    date_str = date_obj.strftime('%Y-%m-%d')
                except:
                    date_str = recorded_at[:10] if len(recorded_at) >= 10 else datetime.now().strftime('%Y-%m-%d')
            else:
                date_str = datetime.now().strftime('%Y-%m-%d')
            
            member_count = stat.get('member_count', 0) or 0
            
            # محاسبه تغییر نسبت به دیروز
            change_from_yesterday = 0
            if channel_id:
                yesterday_stats = self.db.get_yesterday_stats(channel_id)
                if yesterday_stats:
                    change_from_yesterday = member_count - (yesterday_stats.get('member_count', 0) or 0)
            
            # محاسبه تغییر نسبت به اولین روز
            change_from_first_day = 0
            if channel_id:
                first_stats = self.db.get_first_stats(channel_id)
                if first_stats:
                    change_from_first_day = member_count - (first_stats.get('member_count', 0) or 0)
            
            positive_change = stat.get('positive_change', 0) or 0
            
            # فرمت username_date
            date_parts = date_str.split('-')
            if len(date_parts) == 3:
                day = str(int(date_parts[2]))
                date_formatted = f"{date_parts[0]}-{date_parts[1]}-{day}"
            else:
                date_formatted = date_str
            
            # نمایش previous_telegram_id اگر موجود باشد
            previous_id = stat.get('previous_telegram_id')
            if previous_id:
                username_date = f"{username} {date_formatted} (قبلی: {previous_id})"
            else:
                username_date = f"{username} {date_formatted}"
            
            row = [
                title,
                username_date,
                member_count,
                change_from_yesterday,
                change_from_first_day,
                positive_change
            ]
            ws.append(row)
        
        # تنظیم عرض ستون‌ها
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        
        # ذخیره فایل
        filename = f"channel_stats_{category}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    async def handle_text_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """مدیریت پیام‌های متنی"""
        if not await self.check_admin(update, context):
            return
        
        message = self.get_message(update)
        if not message or not message.text:
            return
        
        text = message.text.strip()
        
        # فقط دکمه‌های کیبورد اصلی را پردازش کن
        if text == "📋 لیست کانال‌ها":
            await self.list_channels(update, context)
        elif text == "📊 مشاهده آمار":
            await self.show_stats(update, context)
        elif text == "➕ افزودن کانال":
            await self.add_channel_start(update, context)
        elif text == "🗑️ حذف کانال":
            await self.remove_channel_start(update, context)
        elif text == "📁 ایجاد دسته‌بندی":
            await self.create_category_start(update, context)
        elif text == "🗑️ حذف دسته‌بندی":
            await self.delete_category_start(update, context)
        elif text == "📋 لیست دسته‌بندی‌ها":
            await self.list_categories(update, context)
        elif text == "📥 خروجی اکسل":
            await self.export_excel(update, context)
        elif text == "🔄 Reset آمار":
            await self.reset_stats(update, context)
        elif text == "⚡ بررسی فوری":
            user_id = update.effective_user.id
            # محاسبه زمان تقریبی بر اساس تعداد کانال‌ها (همه کانال‌های فعال)
            channels = self.db.get_all_active_channels()
            channel_count = len(channels)
            # هر کانال حدود 2-3 ثانیه زمان می‌برد
            estimated_seconds = channel_count * 3
            estimated_minutes = estimated_seconds // 60
            estimated_secs = estimated_seconds % 60
            
            time_text = ""
            if estimated_minutes > 0:
                time_text = f"حدود {estimated_minutes} دقیقه"
                if estimated_secs > 0:
                    time_text += f" و {estimated_secs} ثانیه"
            else:
                time_text = f"حدود {estimated_seconds} ثانیه"
            
            success = self.trigger_immediate_check(user_id)
            if success:
                await message.reply_text(
                    "⚡ درخواست بررسی فوری ارسال شد!\n\n"
                    f"⏱️ زمان تقریبی: {time_text}\n"
                    f"📊 تعداد کانال‌ها: {channel_count}\n\n"
                    "به محض اتمام، به شما اطلاع داده می‌شود.",
                    reply_markup=self.get_main_keyboard()
                )
            else:
                await message.reply_text(
                    "❌ خطا در ارسال درخواست بررسی فوری!",
                    reply_markup=self.get_main_keyboard()
                )
    
    async def handle_callback_query(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """مدیریت callback query"""
        query = update.callback_query
        if not query:
            return
        
        data = query.data
        
        if data == "list_channels":
            await query.answer("در حال دریافت لیست کانال‌ها...")
            channels = self.db.get_all_active_channels()  # همه کانال‌های فعال را بگیر
            
            if not channels:
                await query.edit_message_text(
                    "📭 هیچ کانال فعالی وجود ندارد.\n\n"
                    "از دکمه ➕ افزودن کانال برای افزودن کانال جدید استفاده کنید.",
                    reply_markup=self.get_inline_keyboard()
                )
                return
            
            text = "📋 لیست کانال‌ها:\n\n"
            for i, channel in enumerate(channels, 1):
                title = channel.get('title', 'بدون عنوان')
                username = channel['username']
                category = channel.get('category', '')
                is_member = channel.get('is_member', 0)
                member_status = "✅ عضو" if is_member else "⏳ در انتظار عضویت"
                
                text += f"{i}. {title}\n"
                if username.startswith('http') or username.startswith('+'):
                    text += f"   🔗 {username[:40]}...\n"
                else:
                    text += f"   @{username}\n"
                if category:
                    text += f"   📁 {category}\n"
                text += f"   {member_status}\n\n"
            
            text += f"\n📊 تعداد کل: {len(channels)} کانال"
            
            await query.edit_message_text(text, reply_markup=self.get_inline_keyboard())
        
        elif data == "show_stats":
            await query.answer("در حال دریافت آمار...")
            stats = self.db.get_all_stats()
            
            if not stats:
                await query.edit_message_text(
                    "📊 هیچ آماری ثبت نشده است.\n\n"
                    "ربات رصد باید در حال اجرا باشد تا آمار ثبت شود.",
                    reply_markup=self.get_inline_keyboard()
                )
                return
            
            # گروه‌بندی آمار بر اساس دسته‌بندی
            stats_by_category = {}
            for stat in stats:
                category = stat.get('category') or 'بدون دسته‌بندی'
                if category not in stats_by_category:
                    stats_by_category[category] = []
                stats_by_category[category].append(stat)
            
            text = "📊 آمار کانال‌ها\n"
            text += "─" * 30 + "\n\n"
            
            total_channels = sum(len(stats) for stats in stats_by_category.values())
            text += f"📈 تعداد کل: {total_channels} کانال\n"
            text += f"📁 دسته‌بندی‌ها: {len(stats_by_category)}\n\n"
            
            # برای callback query، فقط 3 دسته اول و 3 کانال اول هر دسته را نشان می‌دهیم
            category_count = 0
            for category, category_stats in sorted(stats_by_category.items()):
                if category_count >= 3:
                    remaining = len(stats_by_category) - category_count
                    text += f"\n... و {remaining} دسته‌بندی دیگر"
                    break
                
                text += f"┌─ 📁 {category}\n"
                text += f"│  ({len(category_stats)} کانال)\n"
                text += f"├" + "─" * 28 + "\n"
                
                for i, stat in enumerate(category_stats[:3], 1):
                    channel_id = stat.get('id')
                    username = stat.get('username', 'نامشخص')
                    title = stat.get('title', 'بدون عنوان')
                    member_count = stat.get('member_count', 0) or 0
                    member_change = stat.get('member_change', 0) or 0
                    recorded_at = stat.get('recorded_at', '')
                    
                    # محاسبه تغییرات
                    change_from_first = 0
                    if channel_id:
                        first_stats = self.db.get_first_stats(channel_id)
                        if first_stats and first_stats.get('member_count'):
                            change_from_first = member_count - (first_stats.get('member_count', 0) or 0)
                    
                    # نمایش username
                    if username.startswith('http') or username.startswith('+') or username.startswith('t.me'):
                        username_display = "🔗 لینک پرایوت"
                    else:
                        username_display = f"@{username}"
                    
                    status_icon = "📈" if member_change > 0 else "📉" if member_change < 0 else "➡️"
                    
                    text += f"│\n"
                    text += f"│ {i}. {title}\n"
                    text += f"│    {username_display}\n"
                    text += f"│    👥 {member_count:,}\n"
                    
                    if member_change != 0:
                        sign = "+" if member_change > 0 else ""
                        text += f"│    {status_icon} {sign}{member_change:,}\n"
                    
                    if change_from_first != 0:
                        first_sign = "+" if change_from_first > 0 else ""
                        text += f"│    📊 {first_sign}{change_from_first:,}\n"
                
                if len(category_stats) > 3:
                    text += f"│\n│    ... و {len(category_stats) - 3} کانال دیگر\n"
                
                text += f"└" + "─" * 28 + "\n\n"
                category_count += 1
            
            await query.edit_message_text(text, reply_markup=self.get_inline_keyboard())
        
        elif data == "add_channel":
            await query.answer()
            await query.edit_message_text(
                "➕ افزودن کانال جدید\n\n"
                "لطفاً یوزرنیم کانال را ارسال کنید:\n"
                "مثال: channel_name\n"
                "یا: @channel_name\n\n"
                "❌ برای لغو /cancel را بفرستید"
            )
            # تنظیم state برای ConversationHandler
            context.user_data['waiting_channel'] = True
        
        elif data == "remove_channel":
            await query.answer()
            await self.remove_channel_start(update, context)
        
        elif data == "export_excel":
            await query.answer()
            # نمایش منوی انتخاب نوع خروجی
            # دریافت لیست همه دسته‌بندی‌های موجود
            categories = self.db.get_all_categories()
            
            if not categories:
                await query.edit_message_text(
                    "❌ هیچ دسته‌بندی موجودی برای خروجی وجود ندارد!\n\n"
                    "لطفاً ابتدا دسته‌بندی ایجاد کنید.",
                    reply_markup=self.get_inline_keyboard()
                )
                return
            
            # ساخت کیبورد اینلاین
            keyboard = [
                [InlineKeyboardButton("📊 خروجی همه دسته‌ها", callback_data="export_all_categories")]
            ]
            
            # اضافه کردن دکمه‌های دسته‌بندی‌ها
            for category in sorted(categories):
                keyboard.append([InlineKeyboardButton(f"📁 {category}", callback_data=f"export_category_{category}")])
            
            keyboard.append([InlineKeyboardButton("❌ لغو", callback_data="cancel_export")])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                "📥 انتخاب نوع خروجی:\n\n"
                "لطفاً نوع خروجی مورد نظر را انتخاب کنید:",
                reply_markup=reply_markup
            )
        
        elif data == "export_all_categories":
            await self.export_excel_all(update, context)
        
        elif data.startswith("export_category_"):
            category = data.replace("export_category_", "")
            await self.export_excel_category(update, context, category)
        
        elif data == "cancel_export":
            await query.answer("لغو شد")
            await query.edit_message_text(
                "❌ عملیات لغو شد.",
                reply_markup=self.get_inline_keyboard()
            )
        
        elif data.startswith("delete_category_"):
            category = data.replace("delete_category_", "")
            await query.answer()
            
            # حذف دسته‌بندی (تبدیل همه کانال‌های آن به NULL)
            success = self.db.delete_category(category)
            
            if success:
                await query.edit_message_text(
                    f"✅ دسته‌بندی '{category}' با موفقیت حذف شد!\n\n"
                    f"همه کانال‌های این دسته‌بندی به 'بدون دسته‌بندی' تغییر یافتند.",
                    reply_markup=self.get_inline_keyboard()
                )
            else:
                await query.edit_message_text(
                    f"❌ خطا در حذف دسته‌بندی!",
                    reply_markup=self.get_inline_keyboard()
                )
        
        elif data == "cancel_delete_category":
            await query.answer("لغو شد")
            await query.edit_message_text(
                "❌ عملیات لغو شد.",
                reply_markup=self.get_inline_keyboard()
            )
        
        elif data == "confirm_reset_all":
            await query.answer("در حال Reset آمار...")
            success = self.db.reset_channel_stats()
            
            if success:
                await query.edit_message_text(
                    "✅ آمار با موفقیت Reset شد!\n\n"
                    "تمام تغییرات (member_change, views_change, ...) صفر شدند.\n"
                    "تعداد فعلی اعضا و اطلاعات کانال‌ها حفظ شدند.",
                    reply_markup=self.get_inline_keyboard()
                )
            else:
                await query.edit_message_text(
                    "❌ خطا در Reset آمار!",
                    reply_markup=self.get_inline_keyboard()
                )
        
        elif data == "cancel_reset":
            await query.answer("لغو شد")
            await query.edit_message_text(
                "❌ عملیات Reset لغو شد.",
                reply_markup=self.get_inline_keyboard()
            )
        
        elif data == "trigger_check":
            await query.answer("درخواست بررسی فوری ارسال شد...")
            user_id = update.effective_user.id
            success = self.trigger_immediate_check(user_id)
            if success:
                await query.edit_message_text(
                    "⚡ درخواست بررسی فوری ارسال شد!\n\n"
                    "ربات رصد در حال بررسی کانال‌ها است...\n"
                    "به محض اتمام، به شما اطلاع داده می‌شود.",
                    reply_markup=self.get_inline_keyboard()
                )
            else:
                await query.edit_message_text(
                    "❌ خطا در ارسال درخواست بررسی فوری!",
                    reply_markup=self.get_inline_keyboard()
                )
        
        elif data == "cancel_remove":
            await query.answer()
            await query.edit_message_text(
                "❌ عملیات لغو شد.",
                reply_markup=self.get_inline_keyboard()
            )
        
        elif data.startswith("remove_"):
            await query.answer()
            try:
                channel_id = int(data.split("_")[1])
                await self.remove_channel_confirm(update, context, channel_id)
            except Exception as e:
                await query.answer(f"❌ خطا: {e}", show_alert=True)
    
    def run(self):
        """اجرای ربات"""
        token = self.config.get('bot_token')
        
        if not token:
            token = input("لطفاً توکن ربات تلگرام را وارد کنید: ").strip()
            self.config['bot_token'] = token
            self.save_config()
        
        application = Application.builder().token(token).build()
        
        # Conversation handler برای افزودن کانال
        add_channel_conv = ConversationHandler(
            entry_points=[
                CommandHandler("add", self.add_channel_start),
                MessageHandler(filters.Regex("^➕ افزودن کانال$"), self.add_channel_start)
            ],
            states={
                self.WAITING_CHANNEL_USERNAME: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.add_channel_process)
                ],
                self.WAITING_CHANNEL_CATEGORY: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.add_channel_category)
                ],
            },
            fallbacks=[CommandHandler("cancel", self.cancel)],
        )
        
        # Conversation handler برای ایجاد دسته‌بندی
        create_category_conv = ConversationHandler(
            entry_points=[
                MessageHandler(filters.Regex("^📁 ایجاد دسته‌بندی$"), self.create_category_start)
            ],
            states={
                self.WAITING_CATEGORY_NAME: [
                    MessageHandler(filters.TEXT & ~filters.COMMAND, self.create_category_process)
                ],
            },
            fallbacks=[CommandHandler("cancel", self.cancel)],
        )
        
        # ثبت handlerها - ترتیب مهم است!
        application.add_handler(CommandHandler("start", self.start))
        application.add_handler(CommandHandler("cancel", self.cancel))
        application.add_handler(add_channel_conv)  # باید قبل از MessageHandler باشد
        application.add_handler(create_category_conv)  # باید قبل از MessageHandler باشد
        application.add_handler(CallbackQueryHandler(self.handle_callback_query))
        
        # دستورات قدیمی برای سازگاری
        application.add_handler(CommandHandler("channels", self.list_channels))
        application.add_handler(CommandHandler("stats", self.show_stats))
        application.add_handler(CommandHandler("remove", self.remove_channel_start))
        application.add_handler(CommandHandler("export", self.export_excel))
        
        # در آخر MessageHandler را اضافه می‌کنیم
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, self.handle_text_message))
        
        # اضافه کردن JobQueue برای چک کردن notification
        job_queue = application.job_queue
        if job_queue:
            # چک کردن notification هر 3 ثانیه
            job_queue.run_repeating(
                self.check_and_notify,
                interval=3.0,
                first=3.0
            )
            print("✅ JobQueue برای چک کردن notification فعال شد")
        else:
            print("⚠️ JobQueue در دسترس نیست!")
        
        print("🤖 ربات مدیریتی شروع به کار کرد...")
        application.run_polling(allowed_updates=Update.ALL_TYPES)


def main():
    bot = AdminBot()
    bot.run()


if __name__ == "__main__":
    main()

